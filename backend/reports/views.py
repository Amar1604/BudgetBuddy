from rest_framework import viewsets, permissions, status
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from django.utils import timezone
import datetime
import csv
import calendar
from django.http import HttpResponse

from income.models import Income
from expenses.models import Expense
from budgets.models import Budget
from savings.models import SavingsGoal
from notifications.models import Notification
from .models import Report
from .serializers import ReportSerializer


def generate_report_data(user, report_type, start_date, end_date):
    if report_type == 'income_summary':
        incomes = Income.objects.filter(user=user, income_date__range=(start_date, end_date))
        total_income = incomes.aggregate(total=Sum('amount'))['total'] or 0.0
        
        grouped = {}
        for source, label in Income.INCOME_SOURCES:
            amount = incomes.filter(source=source).aggregate(total=Sum('amount'))['total'] or 0.0
            if amount > 0:
                grouped[source] = float(amount)
                
        return {
            "total_income": float(total_income),
            "by_source": grouped,
            "count": incomes.count()
        }
        
    elif report_type == 'expense_summary':
        expenses = Expense.objects.filter(user=user, expense_date__range=(start_date, end_date))
        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0.0
        
        grouped = {}
        for cat, label in Expense.EXPENSE_CATEGORIES:
            amount = expenses.filter(category=cat).aggregate(total=Sum('amount'))['total'] or 0.0
            if amount > 0:
                grouped[cat] = float(amount)
                
        return {
            "total_expense": float(total_expense),
            "by_category": grouped,
            "count": expenses.count()
        }
        
    elif report_type == 'budget_vs_actual':
        expenses = Expense.objects.filter(user=user, expense_date__range=(start_date, end_date))
        
        # Find all months between start_date and end_date
        months_years = set()
        import calendar
        curr = start_date.replace(day=1)
        while curr <= end_date:
            months_years.add((curr.month, curr.year))
            if curr.month == 12:
                curr = curr.replace(year=curr.year + 1, month=1)
            else:
                curr = curr.replace(month=curr.month + 1)
                
        category_budgets = {}
        for m, y in months_years:
            buds = Budget.objects.filter(user=user, month=m, year=y)
            for b in buds:
                category_budgets[b.category] = category_budgets.get(b.category, 0.0) + float(b.budget_amount)
                
        category_expenses = {}
        for cat, label in Expense.EXPENSE_CATEGORIES:
            amount = expenses.filter(category=cat).aggregate(total=Sum('amount'))['total'] or 0.0
            if amount > 0 or cat in category_budgets:
                category_expenses[cat] = float(amount)
                
        comparisons = {}
        all_cats = set(list(category_budgets.keys()) + list(category_expenses.keys()))
        for cat in all_cats:
            b_val = category_budgets.get(cat, 0.0)
            a_val = category_expenses.get(cat, 0.0)
            comparisons[cat] = {
                "budget": b_val,
                "actual": a_val,
                "variance": b_val - a_val
            }
            
        return {
            "categories": comparisons,
            "total_budget": sum(category_budgets.values()),
            "total_actual": sum(category_expenses.values())
        }
        
    elif report_type == 'net_worth' or report_type == 'custom':
        incomes = Income.objects.filter(user=user, income_date__range=(start_date, end_date))
        expenses = Expense.objects.filter(user=user, expense_date__range=(start_date, end_date))
        total_income = incomes.aggregate(total=Sum('amount'))['total'] or 0.0
        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0.0
        
        return {
            "total_income": float(total_income),
            "total_expense": float(total_expense),
            "net_savings": float(total_income) - float(total_expense)
        }
        
    return {}


class ReportViewSet(viewsets.ModelViewSet):
    serializer_class = ReportSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        return Report.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        user = self.request.user
        validated_data = serializer.validated_data
        report_type = validated_data.get('report_type')
        start_date = validated_data.get('date_range_start')
        end_date = validated_data.get('date_range_end')
        
        calculated_data = generate_report_data(user, report_type, start_date, end_date)
        serializer.save(user=user, data=calculated_data)

    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        report = self.get_object()
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_bold = Font(name='Arial', size=10, bold=True)
        font_regular = Font(name='Arial', size=10)
        
        fill_title = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        fill_summary = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        thin = Side(border_style="thin", color="D1D5DB")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title Block
        ws.merge_cells('A1:D2')
        title_cell = ws['A1']
        title_cell.value = f"BudgetBuddy - {report.title}"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Metadata
        ws['A4'] = "Report Type:"
        ws['A4'].font = font_bold
        ws['B4'] = report.get_report_type_display()
        ws['B4'].font = font_regular
        
        ws['C4'] = "Generated At:"
        ws['C4'].font = font_bold
        ws['D4'] = report.generated_at.strftime('%Y-%m-%d %H:%M')
        ws['D4'].font = font_regular
        
        ws['A5'] = "Date Range:"
        ws['A5'].font = font_bold
        ws['B5'] = f"{report.date_range_start} to {report.date_range_end}"
        ws['B5'].font = font_regular
        
        current_row = 7
        data = report.data or {}
        
        if report.report_type == 'income_summary':
            ws.cell(row=current_row, column=1, value="Income Source").font = font_header
            ws.cell(row=current_row, column=1).fill = fill_header
            ws.cell(row=current_row, column=2, value="Amount").font = font_header
            ws.cell(row=current_row, column=2).fill = fill_header
            current_row += 1
            
            by_source = data.get('by_source', {})
            for source, amt in by_source.items():
                ws.cell(row=current_row, column=1, value=source.replace('_', ' ').title()).font = font_regular
                c_amt = ws.cell(row=current_row, column=2, value=amt)
                c_amt.font = font_regular
                c_amt.number_format = '"₹"#,##0.00'
                current_row += 1
                
            ws.cell(row=current_row, column=1, value="Total Income").font = font_bold
            ws.cell(row=current_row, column=1).fill = fill_summary
            c_tot = ws.cell(row=current_row, column=2, value=data.get('total_income', 0.0))
            c_tot.font = font_bold
            c_tot.fill = fill_summary
            c_tot.number_format = '"₹"#,##0.00'
            
        elif report.report_type == 'expense_summary':
            ws.cell(row=current_row, column=1, value="Expense Category").font = font_header
            ws.cell(row=current_row, column=1).fill = fill_header
            ws.cell(row=current_row, column=2, value="Amount").font = font_header
            ws.cell(row=current_row, column=2).fill = fill_header
            current_row += 1
            
            by_category = data.get('by_category', {})
            for category, amt in by_category.items():
                ws.cell(row=current_row, column=1, value=category.replace('_', ' ').title()).font = font_regular
                c_amt = ws.cell(row=current_row, column=2, value=amt)
                c_amt.font = font_regular
                c_amt.number_format = '"₹"#,##0.00'
                current_row += 1
                
            ws.cell(row=current_row, column=1, value="Total Expense").font = font_bold
            ws.cell(row=current_row, column=1).fill = fill_summary
            c_tot = ws.cell(row=current_row, column=2, value=data.get('total_expense', 0.0))
            c_tot.font = font_bold
            c_tot.fill = fill_summary
            c_tot.number_format = '"₹"#,##0.00'
            
        elif report.report_type == 'budget_vs_actual':
            headers = ["Category", "Budget", "Actual", "Variance"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
            current_row += 1
            
            categories = data.get('categories', {})
            for cat, vals in categories.items():
                ws.cell(row=current_row, column=1, value=cat.replace('_', ' ').title()).font = font_regular
                
                c_bud = ws.cell(row=current_row, column=2, value=vals.get('budget', 0.0))
                c_bud.font = font_regular
                c_bud.number_format = '"₹"#,##0.00'
                
                c_act = ws.cell(row=current_row, column=3, value=vals.get('actual', 0.0))
                c_act.font = font_regular
                c_act.number_format = '"₹"#,##0.00'
                
                c_var = ws.cell(row=current_row, column=4, value=vals.get('variance', 0.0))
                c_var.font = font_regular
                c_var.number_format = '"₹"#,##0.00'
                current_row += 1
                
            ws.cell(row=current_row, column=1, value="Total").font = font_bold
            ws.cell(row=current_row, column=1).fill = fill_summary
            
            c_tbud = ws.cell(row=current_row, column=2, value=data.get('total_budget', 0.0))
            c_tbud.font = font_bold
            c_tbud.fill = fill_summary
            c_tbud.number_format = '"₹"#,##0.00'
            
            c_tact = ws.cell(row=current_row, column=3, value=data.get('total_actual', 0.0))
            c_tact.font = font_bold
            c_tact.fill = fill_summary
            c_tact.number_format = '"₹"#,##0.00'
            
            c_tvar = ws.cell(row=current_row, column=4, value=data.get('total_budget', 0.0) - data.get('total_actual', 0.0))
            c_tvar.font = font_bold
            c_tvar.fill = fill_summary
            c_tvar.number_format = '"₹"#,##0.00'
            
        else:
            ws.cell(row=current_row, column=1, value="Metric").font = font_header
            ws.cell(row=current_row, column=1).fill = fill_header
            ws.cell(row=current_row, column=2, value="Amount").font = font_header
            ws.cell(row=current_row, column=2).fill = fill_header
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Total Income").font = font_regular
            c_inc = ws.cell(row=current_row, column=2, value=data.get('total_income', 0.0))
            c_inc.font = font_regular
            c_inc.number_format = '"₹"#,##0.00'
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Total Expense").font = font_regular
            c_exp = ws.cell(row=current_row, column=2, value=data.get('total_expense', 0.0))
            c_exp.font = font_regular
            c_exp.number_format = '"₹"#,##0.00'
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Net Savings").font = font_bold
            ws.cell(row=current_row, column=1).fill = fill_summary
            c_net = ws.cell(row=current_row, column=2, value=data.get('net_savings', 0.0))
            c_net.font = font_bold
            c_net.fill = fill_summary
            c_net.number_format = '"₹"#,##0.00'

        for r_idx in range(7, current_row + 1):
            for col_idx in range(1, 5):
                cell = ws.cell(row=r_idx, column=col_idx)
                if cell.value is not None or cell.fill.fill_type is not None:
                    cell.border = border_all
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format == '"₹"#,##0.00':
                    val += '   '
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="BudgetBuddy_{report.id}.xlsx"'
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        report = self.get_object()
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = f'attachment; filename="BudgetBuddy_{report.id}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            textColor=colors.HexColor('#1E3A8A'),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#4B5563'),
            spaceAfter=20
        )
        h2_style = ParagraphStyle(
            'H2Style',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor('#1E3A8A'),
            spaceBefore=15,
            spaceAfter=10
        )
        cell_bold = ParagraphStyle(
            'CellBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#111827')
        )
        cell_regular = ParagraphStyle(
            'CellRegular',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#374151')
        )
        cell_header = ParagraphStyle(
            'CellHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.white
        )

        story.append(Paragraph("BudgetBuddy Financial Statement", title_style))
        story.append(Paragraph(f"Report: {report.title}  |  Range: {report.date_range_start} to {report.date_range_end}", subtitle_style))
        story.append(Spacer(1, 10))
        
        data = report.data or {}
        
        overview_data = [
            [Paragraph("Report Type", cell_bold), Paragraph(report.get_report_type_display(), cell_regular)],
            [Paragraph("Generated At", cell_bold), Paragraph(report.generated_at.strftime('%Y-%m-%d %H:%M'), cell_regular)],
        ]
        
        t_overview = Table(overview_data, colWidths=[120, 420])
        t_overview.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F3F4F6')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
            ('PADDING', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(t_overview)
        story.append(Spacer(1, 20))
        
        table_rows = []
        col_widths = []
        
        if report.report_type == 'income_summary':
            story.append(Paragraph("Income Source Allocation", h2_style))
            table_rows.append([Paragraph("Income Source", cell_header), Paragraph("Amount", cell_header)])
            col_widths = [360, 180]
            
            by_source = data.get('by_source', {})
            for source, amt in by_source.items():
                table_rows.append([
                    Paragraph(source.replace('_', ' ').title(), cell_regular),
                    Paragraph(f"Rs. {amt:,.2f}", cell_regular)
                ])
            table_rows.append([
                Paragraph("Total Income", cell_bold),
                Paragraph(f"Rs. {data.get('total_income', 0.0):,.2f}", cell_bold)
            ])
            
        elif report.report_type == 'expense_summary':
            story.append(Paragraph("Expense Category Allocation", h2_style))
            table_rows.append([Paragraph("Expense Category", cell_header), Paragraph("Amount", cell_header)])
            col_widths = [360, 180]
            
            by_category = data.get('by_category', {})
            for category, amt in by_category.items():
                table_rows.append([
                    Paragraph(category.replace('_', ' ').title(), cell_regular),
                    Paragraph(f"Rs. {amt:,.2f}", cell_regular)
                ])
            table_rows.append([
                Paragraph("Total Expense", cell_bold),
                Paragraph(f"Rs. {data.get('total_expense', 0.0):,.2f}", cell_bold)
            ])
            
        elif report.report_type == 'budget_vs_actual':
            story.append(Paragraph("Budget vs Actual Comparisons", h2_style))
            table_rows.append([
                Paragraph("Category", cell_header),
                Paragraph("Budget", cell_header),
                Paragraph("Actual", cell_header),
                Paragraph("Variance", cell_header)
            ])
            col_widths = [240, 100, 100, 100]
            
            categories = data.get('categories', {})
            for cat, vals in categories.items():
                table_rows.append([
                    Paragraph(cat.replace('_', ' ').title(), cell_regular),
                    Paragraph(f"Rs. {vals.get('budget', 0.0):,.2f}", cell_regular),
                    Paragraph(f"Rs. {vals.get('actual', 0.0):,.2f}", cell_regular),
                    Paragraph(f"Rs. {vals.get('variance', 0.0):,.2f}", cell_regular)
                ])
            table_rows.append([
                Paragraph("Total", cell_bold),
                Paragraph(f"Rs. {data.get('total_budget', 0.0):,.2f}", cell_bold),
                Paragraph(f"Rs. {data.get('total_actual', 0.0):,.2f}", cell_bold),
                Paragraph(f"Rs. {(data.get('total_budget', 0.0) - data.get('total_actual', 0.0)):,.2f}", cell_bold)
            ])
            
        else:
            story.append(Paragraph("Financial Summary Analysis", h2_style))
            table_rows.append([Paragraph("Metric Summary", cell_header), Paragraph("Amount", cell_header)])
            col_widths = [360, 180]
            
            table_rows.append([Paragraph("Total Income", cell_regular), Paragraph(f"Rs. {data.get('total_income', 0.0):,.2f}", cell_regular)])
            table_rows.append([Paragraph("Total Expense", cell_regular), Paragraph(f"Rs. {data.get('total_expense', 0.0):,.2f}", cell_regular)])
            table_rows.append([Paragraph("Net Savings", cell_bold), Paragraph(f"Rs. {data.get('net_savings', 0.0):,.2f}", cell_bold)])

        t_data = Table(table_rows, colWidths=col_widths)
        t_styles = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3B82F6')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('PADDING', (0,0), (-1,-1), 6),
        ]
        
        for idx in range(1, len(table_rows) - 1):
            if idx % 2 == 0:
                t_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#F9FAFB')))
                
        t_styles.append(('BACKGROUND', (0, len(table_rows)-1), (-1, len(table_rows)-1), colors.HexColor('#F3F4F6')))
        t_styles.append(('LINEABOVE', (0, len(table_rows)-1), (-1, len(table_rows)-1), 1.5, colors.HexColor('#9CA3AF')))
        
        t_data.setStyle(TableStyle(t_styles))
        story.append(t_data)
        
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=8,
            textColor=colors.HexColor('#9CA3AF'),
            alignment=1
        )
        story.append(Paragraph("Generated securely via BudgetBuddy platform. All rights reserved.", footer_style))
        
        doc.build(story)
        return response


class DashboardAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        user = request.user
        now = datetime.datetime.now()
        
        try:
            month = int(request.query_params.get('month', now.month))
            year = int(request.query_params.get('year', now.year))
        except ValueError:
            return Response(
                {"error": "month and year must be integers"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        total_income = Income.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0.00
        total_expense = Expense.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0.00
        current_balance = float(total_income) - float(total_expense)
        
        total_budget = Budget.objects.filter(user=user, month=month, year=year).aggregate(total=Sum('budget_amount'))['total'] or 0.00
        
        month_expense = Expense.objects.filter(
            user=user,
            expense_date__year=year,
            expense_date__month=month
        ).aggregate(total=Sum('amount'))['total'] or 0.00
        remaining_budget = float(total_budget) - float(month_expense)
        
        recent_incomes = Income.objects.filter(user=user).order_by('-income_date', '-id')[:10]
        recent_expenses = Expense.objects.filter(user=user).order_by('-expense_date', '-id')[:10]
        
        transactions = []
        for inc in recent_incomes:
            transactions.append({
                "id": inc.id,
                "type": "income",
                "title": inc.title,
                "amount": float(inc.amount),
                "category": inc.source,
                "date": inc.income_date.strftime('%Y-%m-%d'),
                "created_at": inc.created_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ')
            })
            
        for exp in recent_expenses:
            transactions.append({
                "id": exp.id,
                "type": "expense",
                "title": exp.title,
                "amount": float(exp.amount),
                "category": exp.category,
                "date": exp.expense_date.strftime('%Y-%m-%d'),
                "created_at": exp.created_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ')
            })
            
        transactions.sort(key=lambda x: (x['date'], x['created_at']), reverse=True)
        recent_transactions = transactions[:10]
        
        return Response({
            "total_income": float(total_income),
            "total_expense": float(total_expense),
            "current_balance": current_balance,
            "total_budget": float(total_budget),
            "remaining_budget": remaining_budget,
            "recent_transactions": recent_transactions
        })


def get_date_range(filter_type, start_date_str=None, end_date_str=None):
    from datetime import date
    import calendar
    today = timezone.localdate()
    
    if filter_type == 'current_month':
        start_date = date(today.year, today.month, 1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = date(today.year, today.month, last_day)
    elif filter_type == 'previous_month':
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year
        start_date = date(prev_year, prev_month, 1)
        last_day = calendar.monthrange(prev_year, prev_month)[1]
        end_date = date(prev_year, prev_month, last_day)
    elif filter_type == 'custom':
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else date(2000, 1, 1)
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else date(2100, 12, 31)
        except (ValueError, TypeError):
            start_date = date(today.year, today.month, 1)
            last_day = calendar.monthrange(today.year, today.month)[1]
            end_date = date(today.year, today.month, last_day)
    else:
        start_date = date(today.year, today.month, 1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = date(today.year, today.month, last_day)
        
    return start_date, end_date


class MonthlyFinancialReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        now = timezone.localdate()
        try:
            month = int(request.query_params.get('month', now.month))
            year = int(request.query_params.get('year', now.year))
        except ValueError:
            return Response({"error": "month and year must be integers"}, status=status.HTTP_400_BAD_REQUEST)

        # Monthly Income
        total_income = Income.objects.filter(
            user=user,
            income_date__year=year,
            income_date__month=month
        ).aggregate(total=Sum('amount'))['total'] or 0.00
        total_income = float(total_income)

        # Monthly Expense
        total_expense = Expense.objects.filter(
            user=user,
            expense_date__year=year,
            expense_date__month=month
        ).aggregate(total=Sum('amount'))['total'] or 0.00
        total_expense = float(total_expense)

        current_balance = total_income - total_expense

        total_savings = SavingsGoal.objects.filter(user=user).aggregate(total=Sum('saved_amount'))['total'] or 0.00
        total_savings = float(total_savings)

        total_budget = Budget.objects.filter(user=user, month=month, year=year).aggregate(total=Sum('budget_amount'))['total'] or 0.00
        total_budget = float(total_budget)
        
        remaining_budget = total_budget - total_expense

        return Response({
            "month": month,
            "year": year,
            "total_income": total_income,
            "total_expense": total_expense,
            "current_balance": current_balance,
            "total_savings": total_savings,
            "remaining_budget": remaining_budget
        })


class ExpenseReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        filter_type = request.query_params.get('filter_type', 'current_month')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        start_date, end_date = get_date_range(filter_type, start_date_str, end_date_str)
        expenses = Expense.objects.filter(user=user, expense_date__range=[start_date, end_date]).order_by('-expense_date')

        if request.query_params.get('export') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="expense_report_{start_date}_{end_date}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Expense Title', 'Category', 'Amount', 'Date', 'Description'])
            for exp in expenses:
                writer.writerow([
                    exp.title,
                    exp.get_category_display(),
                    float(exp.amount),
                    exp.expense_date.strftime('%Y-%m-%d'),
                    exp.description or ''
                ])
            return response

        data = []
        for exp in expenses:
            data.append({
                "title": exp.title,
                "category": exp.get_category_display(),
                "amount": float(exp.amount),
                "date": exp.expense_date.strftime('%Y-%m-%d'),
                "description": exp.description or ''
            })
        return Response(data)


class SavingsReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        goals = SavingsGoal.objects.filter(user=user).order_by('target_date')

        data = []
        for g in goals:
            remaining = max(float(g.target_amount - g.saved_amount), 0.0)
            pct = (float(g.saved_amount) / float(g.target_amount) * 100) if g.target_amount > 0 else 0.0
            data.append({
                "goal_name": g.goal_name,
                "target_amount": float(g.target_amount),
                "saved_amount": float(g.saved_amount),
                "remaining_amount": remaining,
                "progress_percentage": round(pct, 2),
                "status": g.get_status_display()
            })
        return Response(data)


class CombinedSummaryReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        filter_type = request.query_params.get('filter_type', 'current_month')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        start_date, end_date = get_date_range(filter_type, start_date_str, end_date_str)

        incomes = Income.objects.filter(user=user, income_date__range=[start_date, end_date]).order_by('-income_date')
        expenses = Expense.objects.filter(user=user, expense_date__range=[start_date, end_date]).order_by('-expense_date')

        total_income = float(incomes.aggregate(total=Sum('amount'))['total'] or 0.00)
        total_expense = float(expenses.aggregate(total=Sum('amount'))['total'] or 0.00)
        current_balance = total_income - total_expense

        total_savings = float(SavingsGoal.objects.filter(user=user).aggregate(total=Sum('saved_amount'))['total'] or 0.00)

        months_years = set()
        curr = start_date.replace(day=1)
        while curr <= end_date:
            months_years.add((curr.month, curr.year))
            if curr.month == 12:
                curr = curr.replace(year=curr.year + 1, month=1)
            else:
                curr = curr.replace(month=curr.month + 1)

        total_budget = 0.0
        budget_details = []
        for m, y in months_years:
            buds = Budget.objects.filter(user=user, month=m, year=y)
            for b in buds:
                b_amount = float(b.budget_amount)
                total_budget += b_amount
                spent = float(Expense.objects.filter(
                    user=user,
                    category=b.category,
                    expense_date__year=y,
                    expense_date__month=m
                ).aggregate(total=Sum('amount'))['total'] or 0.00)
                budget_details.append({
                    "category": b.get_category_display(),
                    "month": b.month,
                    "year": b.year,
                    "budget_amount": b_amount,
                    "actual_spent": spent,
                    "remaining": b_amount - spent
                })

        remaining_budget = total_budget - total_expense

        financial_summary = {
            "total_income": total_income,
            "total_expense": total_expense,
            "current_balance": current_balance,
            "total_savings": total_savings,
            "remaining_budget": remaining_budget,
            "total_budget": total_budget
        }

        choices_dict_exp = dict(Expense.EXPENSE_CATEGORIES)
        category_totals = {}
        for exp in expenses:
            category_totals[exp.category] = category_totals.get(exp.category, 0.0) + float(exp.amount)

        expense_summary_grouped = []
        for cat, amt in category_totals.items():
            pct = (amt / total_expense * 100) if total_expense > 0 else 0.0
            expense_summary_grouped.append({
                "category": choices_dict_exp.get(cat, cat.capitalize()),
                "total_amount": amt,
                "percentage": round(pct, 2)
            })
        expense_summary_grouped.sort(key=lambda x: x['total_amount'], reverse=True)

        choices_dict_inc = dict(Income.INCOME_SOURCES)
        source_totals = {}
        for inc in incomes:
            source_totals[inc.source] = source_totals.get(inc.source, 0.0) + float(inc.amount)

        income_summary_grouped = []
        for src, amt in source_totals.items():
            pct = (amt / total_income * 100) if total_income > 0 else 0.0
            income_summary_grouped.append({
                "source": choices_dict_inc.get(src, src.capitalize()),
                "total_amount": amt,
                "percentage": round(pct, 2)
            })
        income_summary_grouped.sort(key=lambda x: x['total_amount'], reverse=True)

        savings_goals = []
        goals = SavingsGoal.objects.filter(user=user).order_by('target_date')
        for g in goals:
            remaining = max(float(g.target_amount - g.saved_amount), 0.0)
            pct = (float(g.saved_amount) / float(g.target_amount) * 100) if g.target_amount > 0 else 0.0
            savings_goals.append({
                "goal_name": g.goal_name,
                "target_amount": float(g.target_amount),
                "saved_amount": float(g.saved_amount),
                "remaining_amount": remaining,
                "progress_percentage": round(pct, 2),
                "status": g.get_status_display()
            })

        notifications = Notification.objects.filter(user=user).order_by('-created_at')[:5]
        notification_list = []
        for n in notifications:
            notification_list.append({
                "title": n.title,
                "message": n.message,
                "type": n.get_notification_type_display(),
                "priority": n.get_priority_display(),
                "is_read": n.is_read,
                "created_at": n.created_at.strftime('%Y-%m-%d %H:%M')
            })

        if request.query_params.get('export') == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="financial_summary_report_{start_date}_{end_date}.csv"'
            writer = csv.writer(response)
            
            writer.writerow(['--- FINANCIAL SUMMARY REPORT ---'])
            writer.writerow(['Date Range', f'{start_date} to {end_date}'])
            writer.writerow([])
            writer.writerow(['Metric', 'Amount'])
            writer.writerow(['Total Income', total_income])
            writer.writerow(['Total Expense', total_expense])
            writer.writerow(['Net Balance', current_balance])
            writer.writerow(['Total Savings (All Time)', total_savings])
            writer.writerow(['Remaining Budget', remaining_budget])
            writer.writerow([])

            writer.writerow(['--- EXPENSE BY CATEGORY ---'])
            writer.writerow(['Category', 'Total Amount', 'Percentage'])
            for item in expense_summary_grouped:
                writer.writerow([item['category'], item['total_amount'], f"{item['percentage']}%"])
            writer.writerow([])

            writer.writerow(['--- INCOME BY SOURCE ---'])
            writer.writerow(['Source', 'Total Amount', 'Percentage'])
            for item in income_summary_grouped:
                writer.writerow([item['source'], item['total_amount'], f"{item['percentage']}%"])
            writer.writerow([])

            writer.writerow(['--- BUDGET VS ACTUAL ---'])
            writer.writerow(['Category', 'Month/Year', 'Budget Amount', 'Actual Spent', 'Remaining'])
            for b in budget_details:
                writer.writerow([b['category'], f"{b['month']}/{b['year']}", b['budget_amount'], b['actual_spent'], b['remaining']])
            writer.writerow([])

            writer.writerow(['--- SAVINGS GOALS ---'])
            writer.writerow(['Goal Name', 'Target Amount', 'Saved Amount', 'Remaining', 'Progress', 'Status'])
            for s in savings_goals:
                writer.writerow([s['goal_name'], s['target_amount'], s['saved_amount'], s['remaining_amount'], f"{s['progress_percentage']}%", s['status']])
            
            return response

        elif request.query_params.get('export') == 'pdf':
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            response = HttpResponse(content_type="application/pdf")
            response['Content-Disposition'] = f'attachment; filename="financial_statement_{start_date}_{end_date}.pdf"'
            
            doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
            story = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'LiveTitleStyle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=20,
                textColor=colors.HexColor('#1E3A8A'),
                spaceAfter=15
            )
            subtitle_style = ParagraphStyle(
                'LiveSubtitleStyle',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=10,
                textColor=colors.HexColor('#4B5563'),
                spaceAfter=20
            )
            h2_style = ParagraphStyle(
                'LiveH2Style',
                parent=styles['Heading2'],
                fontName='Helvetica-Bold',
                fontSize=14,
                textColor=colors.HexColor('#1E3A8A'),
                spaceBefore=15,
                spaceAfter=10
            )
            cell_bold = ParagraphStyle(
                'LiveCellBold',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=10,
                textColor=colors.HexColor('#111827')
            )
            cell_regular = ParagraphStyle(
                'LiveCellRegular',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=10,
                textColor=colors.HexColor('#374151')
            )
            cell_header = ParagraphStyle(
                'LiveCellHeader',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=10,
                textColor=colors.white
            )

            story.append(Paragraph("BudgetBuddy - Live Statement Report", title_style))
            story.append(Paragraph(f"Range: {start_date} to {end_date} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
            story.append(Spacer(1, 10))
            
            story.append(Paragraph("Financial Summary Overview", h2_style))
            summary_rows = [
                [Paragraph("Total Income", cell_bold), Paragraph(f"Rs. {total_income:,.2f}", cell_regular)],
                [Paragraph("Total Expense", cell_bold), Paragraph(f"Rs. {total_expense:,.2f}", cell_regular)],
                [Paragraph("Net Balance", cell_bold), Paragraph(f"Rs. {current_balance:,.2f}", cell_regular)],
                [Paragraph("Remaining Budget", cell_bold), Paragraph(f"Rs. {remaining_budget:,.2f}", cell_regular)],
            ]
            t_sum = Table(summary_rows, colWidths=[180, 360])
            t_sum.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F3F4F6')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(t_sum)
            story.append(Spacer(1, 15))

            if expense_summary_grouped:
                story.append(Paragraph("Expense Breakdown", h2_style))
                exp_rows = [[Paragraph("Category", cell_header), Paragraph("Amount", cell_header), Paragraph("Percentage", cell_header)]]
                for item in expense_summary_grouped:
                    exp_rows.append([
                        Paragraph(item['category'], cell_regular),
                        Paragraph(f"Rs. {item['total_amount']:,.2f}", cell_regular),
                        Paragraph(f"{item['percentage']}%", cell_regular)
                    ])
                t_exp = Table(exp_rows, colWidths=[240, 150, 150])
                t_exp.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3B82F6')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
                    ('PADDING', (0,0), (-1,-1), 5),
                ]))
                story.append(t_exp)
                story.append(Spacer(1, 15))

            if income_summary_grouped:
                story.append(Paragraph("Income Breakdown", h2_style))
                inc_rows = [[Paragraph("Source", cell_header), Paragraph("Amount", cell_header), Paragraph("Percentage", cell_header)]]
                for item in income_summary_grouped:
                    inc_rows.append([
                        Paragraph(item['source'], cell_regular),
                        Paragraph(f"Rs. {item['total_amount']:,.2f}", cell_regular),
                        Paragraph(f"{item['percentage']}%", cell_regular)
                    ])
                t_inc = Table(inc_rows, colWidths=[240, 150, 150])
                t_inc.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#10B981')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
                    ('PADDING', (0,0), (-1,-1), 5),
                ]))
                story.append(t_inc)

            story.append(Spacer(1, 40))
            story.append(Paragraph("Generated securely via BudgetBuddy platform. All rights reserved.", subtitle_style))
            
            doc.build(story)
            return response

        elif request.query_params.get('export') == 'excel':
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Live Financial Statement"
            ws.views.sheetView[0].showGridLines = True
            
            font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            font_bold = Font(name='Arial', size=10, bold=True)
            font_regular = Font(name='Arial', size=10)
            
            fill_title = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            fill_summary = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            
            thin = Side(border_style="thin", color="D1D5DB")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

            ws.merge_cells('A1:C2')
            title_cell = ws['A1']
            title_cell.value = "BudgetBuddy - Live Statement Report"
            title_cell.font = font_title
            title_cell.fill = fill_title
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws['A4'] = "Date Range:"
            ws['A4'].font = font_bold
            ws['B4'] = f"{start_date} to {end_date}"
            ws['B4'].font = font_regular
            
            ws.cell(row=6, column=1, value="Metric").font = font_header
            ws.cell(row=6, column=1).fill = fill_header
            ws.cell(row=6, column=2, value="Amount").font = font_header
            ws.cell(row=6, column=2).fill = fill_header
            
            ws.cell(row=7, column=1, value="Total Income").font = font_regular
            ws.cell(row=7, column=2, value=total_income).number_format = '"₹"#,##0.00'
            ws.cell(row=8, column=1, value="Total Expense").font = font_regular
            ws.cell(row=8, column=2, value=total_expense).number_format = '"₹"#,##0.00'
            
            ws.cell(row=9, column=1, value="Net Balance").font = font_bold
            ws.cell(row=9, column=1).fill = fill_summary
            ws.cell(row=9, column=2, value=current_balance).font = font_bold
            ws.cell(row=9, column=2).fill = fill_summary
            ws.cell(row=9, column=2).number_format = '"₹"#,##0.00'
            
            ws.cell(row=11, column=1, value="Expense Category").font = font_header
            ws.cell(row=11, column=1).fill = fill_header
            ws.cell(row=11, column=2, value="Amount").font = font_header
            ws.cell(row=11, column=2).fill = fill_header
            ws.cell(row=11, column=3, value="Percentage").font = font_header
            ws.cell(row=11, column=3).fill = fill_header
            
            curr_row = 12
            for item in expense_summary_grouped:
                ws.cell(row=curr_row, column=1, value=item['category']).font = font_regular
                c_val = ws.cell(row=curr_row, column=2, value=item['total_amount'])
                c_val.font = font_regular
                c_val.number_format = '"₹"#,##0.00'
                ws.cell(row=curr_row, column=3, value=f"{item['percentage']}%").font = font_regular
                curr_row += 1
                
            for r in range(6, curr_row):
                for c in range(1, 4):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None or cell.fill.fill_type is not None:
                        cell.border = border_all
                        
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename="financial_statement_{start_date}_{end_date}.xlsx"'
            wb.save(response)
            return response

        return Response({
            "date_range": {
                "start_date": start_date.strftime('%Y-%m-%d'),
                "end_date": end_date.strftime('%Y-%m-%d')
            },
            "financial_summary": financial_summary,
            "expense_summary": expense_summary_grouped,
            "income_summary": income_summary_grouped,
            "budget_summary": budget_details,
            "savings_summary": savings_goals,
            "latest_notifications": notification_list
        })


